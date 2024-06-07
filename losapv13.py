# Version history (see Github for more info)
#   v1.0: Initial commit (3/9/2024)
#   v1.1: Fixed Excel output and add disability column (3/30/2024) 

import sys
import os
import warnings
import uuid
import pandas as pd
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill, Font
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTableView, 
                             QAction, QFileDialog, QTextBrowser,
                             QDialog, QProgressDialog, QMessageBox)
from PyQt5.QtCore import QAbstractTableModel, Qt, QUrl, QSettings

# user-defined classes in external files
from settings_ui import Ui_Settings
from agreement import Ui_Agreement_Dialog

__author__      = "William A Coetzee"
__copyright__   = "Copyright Reserved"
__credits__     = "William A Coetzee"
__license__     = "GNU General Public License v3"
__version__     = "1.0.0"
__maintainer__  = "William A. Coetzee"
__email__       = "william.coetzee@gmail.com"
__status__      = "Production"
__progname__    = "LOSAP Points Calculator"

__demo__               = False
__debugging__          = False
__debuggingiar__       = False
__debuggingepcr__      = False
__debuggingother__     = True
__debuggingsettings__  = False

# supress future warnings
pd.set_option('future.no_silent_downcasting', True)
warnings.simplefilter(action='ignore', category=FutureWarning)

def swap_name_order(df_def):
    # There appear to be two versions of the spreadsheet out there. In some, names are 
    # entered as "Last, First" and in others, names are the the form of "First Last"
    # Let's try to rectify this by assuming that a name field containing a comma is in the
    # corerct form. If not, then we will swap the two name entries and add a comma
    # **TODO: Distribute a spreadsheet with names in the form: "Last, First"
    for i in range(len(df_def)):
        fname = df_def.loc[i, "Member Name"].rstrip()
        if (fname.find(',') > 0):
            df_def.loc[i, "Temp name"] = fname
        else:
            # Reverse first and last names 
            new = fname.rsplit(" ",1) 
            df_def.loc[i, "Temp name"] =  new[1] + ', ' + new[0] 

    # Delete the 'Full name' column and rename the 'Temp name' column 
    df_def = df_def.drop(columns=['Member Name'], axis=1)
    df_def = df_def.rename(columns={"Temp name": "Member Name"})
    return (df_def)

class PandasModel(QAbstractTableModel):
    def __init__(self, data):
        super(PandasModel, self).__init__()
        self._data = data

    def rowCount(self, parent=None):
        return len(self._data.index)

    def columnCount(self, parent=None):
        return self._data.columns.size

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            if role == Qt.DisplayRole:
                return str(self._data.iloc[index.row(), index.column()])
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self._data.columns[section]
        return None

class AboutWindow(QMainWindow):
    def __init__(self):
        super(AboutWindow, self).__init__()
        self.setWindowTitle("About")
        self.setGeometry(100, 100, 600, 400)
        self.central_widget = QTextBrowser()
        self.setCentralWidget(self.central_widget)
        self.central_widget.setHtml("<center><h1><p><p>About </h1>"
                                + "<h2><p><p> LOSAP Points Calculator</h2>"
                                + "<p>Version: " + __version__
                                + "<br/>Author: " + __author__ 
                                + "<br>Email: " + __email__
                                + "<br>Copyright: " + __license__
                                + "</p></center>")

class ManualWindow(QMainWindow):
    def __init__(self):
        super(ManualWindow, self).__init__()
        self.setWindowTitle("Manual")
        self.setGeometry(250, 250, 1200, 1200)
        self.central_widget = QTextBrowser()
        self.setCentralWidget(self.central_widget)
        self.central_widget.setOpenExternalLinks(True)
        self.central_widget.setSource(QUrl.fromLocalFile(os.path.abspath("About LOSAP Points Calculator.htm")))

#class SettingsDialog(QDialog):
    #print("test")


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()

        self.setWindowTitle("LOSAP Points")
        self.setGeometry(100, 100, 1800, 1000)
        
        # Restore saved progam settings
        self.settings = QSettings("WAC", "LOSAP Analyzer")
        #print(self.settings.fileName())        
        try:
            self.resize(self.settings.value('window size'))
            self.move(self.settings.value('window position'))
        except:
            pass   

        self.table_view = QTableView()  
        self.setCentralWidget(self.table_view)

        # variables
        #self.colnames = ["Member Name", "Training", "Drills", "Meetings", "Tour of Duty", 
        #                 "Misc. Activity", "Calls Responded To", "Position Held", "Disability", 
        #                 "Total", "SR_Signup", "SR_Calls", "SR_Total"]
        self.colnames = ["Member Name", "Training", "Drills", "Meetings", "Tour of Duty", 
                         "Misc. Activity", "Calls Responded To", "Position Held", "Disability", 
                         "Total", "SR_Total"]
        self.colnamestoadd = ["Training", "Drills", "Meetings", "Tour of Duty", 
                         "Misc. Activity", "Calls Responded To", "Position Held", "Disability"]
        self.original_df = pd.DataFrame(columns=self.colnames)
        self.df = self.original_df.copy()

        # I am responding
        self.iamr_rows_to_skip = 2  # Skip this number of rows before reading data
        self.iamr_rows_end = 251    # Last row containing data (just before 'Name	Total hours')

        # Member reported spreadsheets (All spreadsheets are present in a single directory)
        self.losap_sheet = 'point tracker'
        self.losap_name_pos = 'D4'       # Position of the person's name 
        self.losap_SR_Signups = 'E7'     # Position of the self-reported signup hours
        self.losap_SR_Calls = 'E8'       # Position of the self-reported call hours
        self.losap_rows_to_skip = 9     # Skip this number of rows before reading data
        
        # Output Excel file
        self.output_file_name = '2024-01'
        self.output_worksheet_name = 'Points Summary'

        # create menus and display the empty table 
        self.create_menu()
        self.update_table()
        
        self.accept_agreement()
        
        # Show 'Ready' in the status bar
        self.statusBar().showMessage("Ready", 0)
    
    def accept_agreement(self):
        reg = self.settings.value('Registration ID')
        
        # If software is not registered, user need to accept agreement 
        if reg is None :
            Agreement = QDialog()
            Agree_ui = Ui_Agreement_Dialog()
            Agree_ui.setupUi(Agreement)
            Agreement.show()
            rsp = Agreement.exec_()
            if rsp == QDialog.Accepted:
                print("accepted")
                reg = (0xf82819a8062e ^ uuid.getnode())
                self.settings.setValue('Registration ID', reg) 
            else:
                msg = "This program cannot be used unless the agreement is accepted" 
                QMessageBox.information(self, "Ivalid use", msg)
                sys.exit()
        if (reg != (0xf82819a8062e ^ uuid.getnode())):
            msg = "You need to accept the user agreement on this computer. \nPlease run the program again" 
            QMessageBox.information(self, "New hardware detected", msg)
            self.settings.remove('Registration ID')
            sys.exit()
            
    def create_menu(self):
        menubar = self.menuBar()

        # The menus to use
        file_menu = menubar.addMenu('File')
        edit_menu = menubar.addMenu('Edit')
        help_menu = menubar.addMenu('Help')

        # File menu
        new_action = QAction('New', self)
        new_action.triggered.connect(self.clear_all)

        import_iamresponding_action = QAction('Import the I am Responding Report (xls)', self)
        import_iamresponding_action.triggered.connect(self.import_iamresponding)

        import_epcr_action = QAction('Import the ePCR Report (csv)', self)
        import_epcr_action.triggered.connect(self.import_epcr)

        import_other_action = QAction('Import Member Self-Reports (xlsx)', self)
        import_other_action.triggered.connect(self.import_other)

        export_action = QAction('Export the Results to Excel (xlsx)', self)
        export_action.triggered.connect(self.export_data)

        exit_action = QAction('Exit', self)
        exit_action.triggered.connect(self.close)

        file_menu.addAction(new_action)        
        file_menu.addAction(import_iamresponding_action)
        file_menu.addAction(import_epcr_action)
        file_menu.addAction(import_other_action)
        file_menu.addAction(export_action)
        file_menu.addAction(exit_action)

        # Edit menu
        clear_action = QAction('Clear', self)
        clear_action.triggered.connect(self.clear_all)

        settings_action = QAction('Settings', self)
        settings_action.triggered.connect(self.open_settings)

        edit_menu.addAction(clear_action)        
        edit_menu.addAction(settings_action)
        
        # Help menu
        about_action = QAction('About', self)
        about_action.triggered.connect(self.open_about)

        manual_action = QAction('How to use', self)
        manual_action.triggered.connect(self.open_manual)

        help_menu.addAction(manual_action)        
        help_menu.addAction(about_action)
                
    def open_about(self):
        self.about_window = AboutWindow()
        self.about_window.show()

    def open_manual(self):
        self.manual_window = ManualWindow()
        self.manual_window.show()
    
    def update_table(self):
        self.model = PandasModel(self.df)
        self.table_view.setModel(self.model)

    def clear_all(self):
        self.df = self.original_df.copy()
        self.update_table()

    def open_settings(self):
        # lines copied from the __main__ section of settings_ui.py
        Settings = QDialog()
        Settings_ui = Ui_Settings()
        Settings_ui.setupUi(Settings)
        Settings.show()
        # added to execute the dialog
        rsp = Settings.exec_()
        
        if rsp == QDialog.Accepted:
            # I am responding settings
            try:
                n = Settings_ui.iar_rows_to_skip_d.text()
                self.iamr_rows_to_skip = int(n)
            except ValueError:
                msg = "Please use a number for IAR rows to skip" \
                    + "\nYou entered {n}"
                QMessageBox.information(self, "Ivalid data", msg)
                self.iamr_rows_to_skip = 2 
            
            try:
                n = Settings_ui.iamr_rows_end_d.text()
                self.iamr_rows_end = int(n)
            except ValueError:
                msg = "Please use a number for IAR \'Stop reading at rows\'" \
                    + "\nYou entered {n}"
                QMessageBox.information(self, "Ivalid data", msg)
 
            try:
                n = Settings_ui.losap_rows_to_skip_d.text()
                self.losap_rows_to_skip = int(n)
            except ValueError:
                msg = "Please use a number for CVAC sreadsheet \'Skip this number of rows\'" \
                    + "\nYou entered {n}"
                QMessageBox.information(self, "Ivalid data", msg)
    
            # Member reported spreadsheets
            self.losap_sheet = Settings_ui.losap_sheet_d.text()
            self.losap_name_pos = Settings_ui.losap_name_pos_d.text()
            self.losap_SR_Signups = Settings_ui.losap_SR_Signups_d.text()
            self.losap_SR_Calls = Settings_ui.losap_SR_Calls_d.text()

            
            # Output Excel file
            self.output_file_name = Settings_ui.output_file_name_d.text()
            self.output_worksheet_name = Settings_ui.output_worksheet_name_d.text()
            
            if (__debuggingsettings__ ):
                print('New settings')
                print("AIR Rows to skip: " + str(self.iamr_rows_to_skip))
                print("AIR read to rows: " + str(self.iamr_rows_end))
                print(self.losap_sheet)
                print(self.losap_name_pos)
                print(self.losap_SR_Signups)
                print(self.losap_SR_Calls)
                print(self.losap_rows_to_skip)
                print(self.output_file_name)
                print(self.output_worksheet_name)
        
        else:
            print("Dialog was closed")  
    
    # save settings
    def closeEvent(self, event):
        self.settings.setValue('window size', self.size())
        self.settings.setValue('window position', self.pos())
        #self.settings.setValue('iamr skip rows', self.iamr_rows_to_skip)
        
        
    # ------------------------------------------------------------------- 
    # Calculate the "Tour of Duty" points from the 'I am responding' data
    #   Read the 'I am responding' exported file (sign-ups)
    #   Skip the first 2 rows, and read until row 251
    
    def import_iamresponding(self):
        
        # Ignore code warnings 
        warnings.simplefilter(action='ignore', category=UserWarning)
        
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", 
                                                   "Excel Files (*.xls)", 
                                                   options=options)
        if file_name:
            try:
                #data = pd.read_excel(file_name, header=self.iamr_rows_to_skip - 1, nrows=self.iamr_rows_end - self.iamr_rows_to_skip + 1)
                #data.rename(columns={"Last name": "Member Name", "Shift hours": "Tour of Duty"}, inplace=True)

                df_iamr = pd.read_excel(file_name, skiprows=self.iamr_rows_to_skip, 
                        nrows=self.iamr_rows_end - self.iamr_rows_to_skip - 1)
                
                # create a new column with combined names: 'Last name, first name'
                df_iamr['Member Name'] = df_iamr['Last name'] + ', ' + df_iamr['First name']
                
                # calculate the aggregate shift hours per person
                df_group = df_iamr.groupby("Member Name")
                df_columns = df_group[["Shift hours"]]
                df_iamr_grouped = df_columns.sum().round(decimals=0).reset_index()
                
                if __debuggingiar__:
                    print(df_iamr_grouped.head(5))
                
                # Calculate LOSAP points
                #   Tour of Duty; 20 points maximum per year
                #   One-half (1/2) point for each 6 hours of scheduled duty
                df_iamr_grouped["Tour of Duty"] = (df_iamr_grouped["Shift hours"]/12).round(2)
                
                if __debuggingiar__:
                    print(df_iamr_grouped.head(5))
                
                # If imposing the 20 points maximum, uncomment the next line
                #df_iamr_grouped["Tour of Duty"] = (df_iamr_grouped["Shift hours"]/12).clip(upper=20)
                
                # Delete the hours column
                df_iamr_grouped = df_iamr_grouped.drop(columns=['Shift hours'], axis=1)
                
                # Fix incorrect full names. e.g. 'Smith, Jon' should be 'Smith, John'
                # *** TODO: Ask tem to fix Jon's name in IAR
                df_iamr_grouped['Member Name'] = df_iamr_grouped['Member Name'].str.replace('Smith, Jon','Smith, John')
                
                # lists the headings and the first 5 entries
                if __debuggingiar__:
                    print(df_iamr_grouped.head(5))
                    print(self.df.head(5))

                ############################################################
                # Merge dataframes based on the member name
                self.df = pd.merge(df_iamr_grouped, self.df, on="Member Name", how="left")
                
                try:
                    self.df = pd.merge(df_iamr_grouped, self.df, on="Member Name", how="left")
                except KeyError:
                    self.df["Member Name"]=df_iamr_grouped['Member Name']
                    self.df["Tour of Duty"]=df_iamr_grouped['Tour of Duty']
                
                # add up all the points
                self.df['Total'] = self.df[self.colnamestoadd].sum(axis=1)
                
                # reorder the columns, sort and replace NAN
                self.df = self.df[self.colnames]
                self.df = self.df.sort_values(by=['Member Name'])
                self.df = self.df.fillna(0)
            
                if __demo__:
                    self.df = self.df.head(15)

                #self.df = data.reindex(columns=self.df.columns)
                self.update_table() 
                self.statusBar().showMessage("I am responsing data imported", 0)
            
            except Exception as e:
                print("Error processesing iar:", e)


    # ------------------------------------------------------------------- 
    # Calculate the "Calls Responded To" points from the 'ePCR' data
    # 0.5 points to each call responded to, with a maximum of 25 points per year

    def import_epcr(self):
        
        # Ignore code warnings 
        warnings.simplefilter(action='ignore', category=UserWarning)
        
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open CSV File", "", "CSV Files (*.csv)", options=options)
        if file_name:
            try:
                df_ePCR = pd.read_csv(file_name)
                
                #rename column
                df_ePCR.rename(columns={"Incident Crew Member Full Name": "Member Name"}, inplace=True)
                
                # Remove double spaces from the name field
                df_ePCR['Member Name'] = \
                    df_ePCR['Member Name'].str.replace(r'  ', ' ', regex=False)
                
                # Reverse first and last names 
                new = df_ePCR['Member Name'].str.rsplit(" ", n = 1, expand = True) 
                df_ePCR["First name"]= new[0] 
                df_ePCR["Last name"]= new[1] 
                
                # Replace member name with combined names: 'Last name, first name'
                df_ePCR['Member Name'] = df_ePCR['Last name'] + ', ' + df_ePCR['First name'] 
                
                # Count the number of calls per person
                df_ePCR_grouped = df_ePCR.groupby("Member Name").size().reset_index(name='Calls Responded To')

                # Now halve it to get the actual points
                df_ePCR_grouped['Calls Responded To'] = df_ePCR_grouped['Calls Responded To']/2
                
                # delete the 'Calls Responded To' column
                self.df = self.df.drop('Calls Responded To', axis=1)
                
                # Merge with existing DataFrame and add new 'Calls Responded To' column
                try:
                    #self.df = pd.merge(self.df, df_ePCR_grouped, on="Member Name", how="left")
                    self.df = pd.merge(self.df, df_ePCR_grouped, how="outer", on=["Member Name", "Member Name"])
                except KeyError:
                    self.df["Member Name"]=df_ePCR_grouped['Member Name']
                    self.df["Calls Responded To"]=df_ePCR_grouped['Calls Responded To']

                if __debuggingepcr__:
                    print(self.df.head(5))
                    print(df_ePCR_grouped.head(5))
                    
                # reorder the columns, sort and replace NAN with zero
                self.df = self.df[self.colnames]
                self.df = self.df.sort_values(by=['Member Name'])
                self.df = self.df.fillna(0)
                
                # add up the points
                self.df['Total'] = self.df[self.colnamestoadd].sum(axis=1)

                if __demo__:
                    self.df = self.df.head(15)                
  
                self.update_table()
                self.statusBar().showMessage("ePCR data imported", 0)
                
            except Exception as e:
                print("Error processing ePCR:", e)

    # ------------------------------------------------------------------- 
    # Read member self-reported spreadsheets (all in a single folder)
    #       skip first number of rows (defined by 'losap_rows_to_skip')
    #
    # Categories to parse are "Training", "Drills", "Meetings", "Misc Activity" 
    # "Tour of Duty", "Calls responded to" and "Positions held" are obtained elsewhere

    def import_other(self):
        
        # Ignore code warnings 
        warnings.simplefilter(action='ignore', category=UserWarning)
        
        
        options = QFileDialog.Options()
        directory = QFileDialog.getExistingDirectory(self, "Select Directory", options=options)
        if directory:
            try:
                files = os.listdir(directory)
                total_files = len(files)
                df_losap   = pd.DataFrame() 
                df_losapSR = pd.DataFrame() 
                i=0
                
                # set up a progress dialog
                progress_dialog = QProgressDialog("Importing Excel files...", "Cancel", 0, total_files, self)
                progress_dialog.setWindowTitle("Import Progress")
                progress_dialog.setWindowModality(Qt.WindowModal)
                progress = 0
                
                for file in files:
                    if file.endswith('.xlsx') and not(file.startswith('~')):
                        file_path = os.path.join(directory, file)
                        #df = pd.read_excel(file_path)
                        #dfs.append(df)
                        if __debuggingother__:
                            print(file_path)
                            
                        workbook = pd.read_excel(open(file_path, 'rb'), 
                                                 sheet_name=self.losap_sheet, 
                                                 skiprows=self.losap_rows_to_skip) 
                        
                        # open the workbook again to get the name of the person 
                        wb = load_workbook(filename = file_path)
                        sheet_range = wb[self.losap_sheet]
                        workbook['Member Name'] = sheet_range[self.losap_name_pos].value

                        #   Read the portion of the spreadsheet that contains self-reported hours
                        df_losapSR.loc[i, ['Member Name']] = str(sheet_range[self.losap_name_pos].value)
                        df_losapSR.loc[i,['SR_Signup']] = sheet_range[self.losap_SR_Signups].value
                        df_losapSR.loc[i,['SR_Calls']]  = sheet_range[self.losap_SR_Calls].value
                        i=i+1
                        
                        # update the progress dialog
                        progress += 1
                        progress_dialog.setValue(progress)
                        if progress_dialog.wasCanceled():
                            break
                       
                        df_losap=pd.concat([df_losap, workbook], sort=False)
                        
                        
                if df_losap.shape[0] > 0:
                    print('starting  to process df_losap')
                    df_losap = df_losap.reset_index()
                    df_losapSR = df_losapSR.reset_index()
                    
                    # replace NAN with zero
                    df_losapSR=df_losapSR.fillna(0)
                    print(df_losapSR)
                    
                    # rename some of the headings
                    df_losap.columns = df_losap.columns.str.replace('Activity \n(not hours & calls)', 'Activity') 
                    df_losap.columns = df_losap.columns.str.replace('time spent \n(in hours)', 'Hours') 
                    df_losap.columns = df_losap.columns.str.replace('Description ', 'Description') 

                    # Swap the first and last names if needed
                    df_losap = swap_name_order(df_losap)
                    df_losapSR = swap_name_order(df_losapSR)
                    
                    # Some members fail to complete the Hours field and Python reads this as NaN
                    # Replace NaN with 1, with the assumption that the event lasted 1 hour
                    df_losap['Hours'] = df_losap['Hours'].fillna(1)
                    
                    # drop some columns
                    df_losap = df_losap.drop(columns=['Date', 'Description', 'Notes/Questions', 
                            'Activity code','index'], axis=1)
                    
                    # Delete rows with an undefined Activity
                    df_losap.dropna(subset=['Activity'], inplace=True)
                    
                    # Split up the data by activity
                    #   Categories to parse are "Training", "Drills", "Meetings", "Misc Activity"
                    df_losap_meetings = df_losap[df_losap['Activity']=='Meetings'].reset_index()
                    df_losap_drills   = df_losap[df_losap['Activity']=='Drills, CMEs'].reset_index()
                    df_losap_training = df_losap[df_losap['Activity']=='Training Course'].reset_index()
                    df_losap_misc     = df_losap[df_losap['Activity']=='Miscellaneous'].reset_index()
                    df_losap_disability = df_losap[df_losap['Activity']=='Disability'].reset_index()
                    
                    #-----------------------------------------------------------
                    # Calculate points
                    # ------------------
                    #   Meetings:   1 point per attendance, irrespective of the meeting duration
                    df_losap_meetings = df_losap_meetings.groupby(['Member Name'])['Hours'].agg('count').reset_index()
                    df_losap_meetings = df_losap_meetings.rename(columns={"Hours": "Meetings"})
                    
                    # ------------------
                    #   Training:   1 point/h with a max of 5 points if less than 20 hours
                    #               1 point/h with a max of 10 points between 20-45 hours
                    #               15 points if more than 45 hours
                    #   Here we will simply calculate the points and not consider annual limits 
                    df_losap_training = df_losap_training.groupby(['Member Name'])['Hours'].agg('sum').reset_index()
                    df_losap_training = df_losap_training.rename(columns={"Hours": "Training"})
                    
                    # ------------------
                    #   Drills:     One (1) point per drill or seminar (minimum two hours duration).
                    #               2 point if more than 4 hours 
                    # df_losap_drills['Drills']=0
                    # for i in df_losap_drills.index:
                    #     h = float(df_losap_drills['Hours'][i])
                    #     if (h >= 2 and h <= 4):
                    #         df_losap_drills.loc[i, "Drills"] = 1
                    #     elif (h > 4):
                    #         df_losap_drills.loc[i, "Drills"] = 2
                    df_losap_drills = df_losap_drills.groupby(['Member Name'])['Points'].agg('sum').reset_index()
                    df_losap_drills = df_losap_drills.rename(columns={"Points": "Drills"})
                    
                    # ------------------
                    #   Misc:     One point per activity for participation in activities 
                    df_losap_misc = df_losap_misc.groupby(['Member Name'])['Points'].agg('sum').reset_index()
                    df_losap_misc = df_losap_misc.rename(columns={"Points": "Misc Activity"})
                    
                    # ------------------
                    #   Disability: Read the points from the points column & cap at 5
                    df_losap_disability = df_losap_disability.groupby(['Member Name'])['Points'].agg('sum').reset_index()
                    df_losap_disability = df_losap_disability.rename(columns={"Points": "Disability"})
                    df_losap_disability['Disability'] = df_losap_disability['Disability'].clip(upper=5.0)
                
                    # ------------------
                    #  Self-reported points for Tour of Duty (signups)
                    #       One-half (1/2) point for each 6 hours of scheduled duty
                    df_losapSR["SR_Signup"] = (df_losapSR["SR_Signup"]/12).round(3)
                    df_losapSR['SR_Signup'] = df_losapSR['SR_Signup'].fillna(0)

                    # ------------------
                    #  Self-reported points for Calls Responded To ['SR Calls Responded To']
                    #  0.5 points to each call responded to, with a maximum of 25 points per year
                    df_losapSR["SR_Calls"] = (df_losapSR["SR_Calls"]/2)
                    df_losapSR['SR_Calls'] = df_losapSR['SR_Calls'].fillna(0)

                    # New section to join the two self-reported columns
                    df_losapSR["SR_Total"] = df_losapSR["SR_Calls"] + df_losapSR['SR_Signup']

                    # Merge meetings with existing DataFrame 
                    self.df = pd.merge(self.df, df_losap_meetings, how="outer", on=["Member Name", "Member Name"])
                    self.df = self.df.drop('Meetings_x', axis=1)
                    self.df = self.df.rename(columns={'Meetings_y': 'Meetings'})
                    
                    # Merge drills with existing DataFrame 
                    self.df = pd.merge(self.df, df_losap_drills, how="outer", on=["Member Name", "Member Name"])
                    self.df = self.df.drop('Drills_x', axis=1)
                    self.df = self.df.rename(columns={'Drills_y': 'Drills'})
                    
                    # Merge training with existing DataFrame 
                    self.df = pd.merge(self.df, df_losap_training, how="outer", on=["Member Name", "Member Name"])
                    self.df = self.df.drop('Training_x', axis=1)
                    self.df = self.df.rename(columns={'Training_y': 'Training'})         
                    
                    # Merge misc with existing DataFrame 
                    self.df = pd.merge(self.df, df_losap_misc, how="outer", on=["Member Name", "Member Name"])
                    self.df = self.df.drop('Misc. Activity', axis=1)
                    self.df = self.df.rename(columns={'Misc Activity': 'Misc. Activity'})      
 
                    # Merge disability with existing DataFrame 
                    #self.df = df_losap_disability.rename(columns={'Disability': 'Disability'})
                    self.df = pd.merge(self.df, df_losap_disability, how="outer", on=["Member Name", "Member Name"])
                    self.df = self.df.drop('Disability_x', axis=1)
                    self.df = self.df.rename(columns={'Disability_y': 'Disability'})  

                    # now that we've calcualted the total, drop the Signup and Calls columns
                    df_losapSR = df_losapSR.drop('SR_Signup', axis=1)
                    df_losapSR = df_losapSR.drop('SR_Calls', axis=1)

                    # Merge self-reported points with df_losapSRDataframe
                    self.df = pd.merge(self.df, df_losapSR, how="outer", on=["Member Name", "Member Name"]) 
                    #self.df = self.df.drop(['SR_Signup_x','SR_Calls_x'], axis=1)
                    #self.df = self.df.rename(columns={'SR_Signup_y': 'SR_Signup'})
                    #self.df = self.df.rename(columns={'SR_Calls_y': 'SR_Calls'})
                    self.df = self.df.drop('SR_Total_x', axis=1)
                    self.df = self.df.rename(columns={'SR_Total_y': 'SR_Total'})

                    
                    
                    # reorder the columns, sort and replace NAN with zero
                    self.df = self.df[self.colnames]
                    self.df = self.df.sort_values(by=['Member Name'])
                    self.df = self.df.fillna(0)
                    
                    # add up the points
                    self.df['Total'] = self.df[self.colnamestoadd].sum(axis=1)
      
                    if __demo__:
                        self.df = self.df.head(15)   
                    
                    progress_dialog.close()
                    self.update_table()
            except Exception as e:
                print("Error processing self-reporting spreadsheets:", e)

    def export_data(self):
        #bk_blue = 4472c4
        #bk_drkblue = 305496
        #cell_gray = d9d9d9
        #cell_blue = d6dce4
        options = QFileDialog.Options()
        default_file_name = self.output_file_name + ' Points Record.xlsx'        
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", default_file_name, 
                                                   "Excel Files (*.xlsx)", options=options)
        if file_name:
            try:
                with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:  
                    self.df.to_excel(writer, sheet_name = self.output_worksheet_name, index = False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets[self.output_worksheet_name]
                    
                    # Define formats for header and alternating rows
                    header_format = workbook.add_format({'bg_color': '#4472c4', 'font_color': 'white'})
                    even_row_format = workbook.add_format({'bg_color': '#d9d9d9'})
                    odd_row_format = workbook.add_format({'bg_color': '#FFFFFF'})
                    
                    # Apply header format to the first row (i.e., the header row)
                    for col_num, value in enumerate(self.df.columns.values):
                        worksheet.write(0, col_num, value, header_format)
                    #worksheet.write(0, col_num+1, value, header_format)
                    worksheet.write(0, col_num, value, header_format)
                    
                    # Apply alternating row formats to the data rows
                    #for row_num in range(1, self.df.shape[0] + 1):
                    for row_num in range(1, self.df.shape[0]):
                        if row_num % 2 == 0:
                            format_to_apply = even_row_format
                        else:
                            format_to_apply = odd_row_format
                        for col_num, value in enumerate(self.df.iloc[row_num - 1]):
                             #worksheet.write(row_num, col_num+1, value, format_to_apply)
                             worksheet.write(row_num, col_num, value, format_to_apply)
                    
                    # Save and close the workbook
                    #writer.save()
                    writer.close()
                    
            except Exception as e:
                print("Error in processing the export file:", e)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
